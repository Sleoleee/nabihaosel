"""
Bangun workbook Excel hasil forecast SES (2 tahun) + grafik.
Output: SES_Forecast_Result.xlsx
"""
import numpy as np
import pandas as pd
from statsmodels.tsa.holtwinters import SimpleExpSmoothing

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

DATA_PATH = "Case_2_FACT.xlsx"
HORIZON = 24

# ----------------------------------------------------------------- compute
df = pd.read_excel(DATA_PATH, header=None, names=["date", "value"])
df["date"] = pd.to_datetime(df["date"])
df = df.sort_values("date").set_index("date")
full = pd.date_range(df.index.min(), df.index.max(), freq="MS")
series = df["value"].reindex(full)
n_missing = int(series.isna().sum())
series = series.interpolate(method="linear")
series.index.freq = "MS"

model = SimpleExpSmoothing(series, initialization_method="estimated").fit()
alpha = float(model.params["smoothing_level"])
fitted = model.fittedvalues
resid = series.values - fitted.values
mape = float(np.mean(np.abs(resid[series.values != 0] / series.values[series.values != 0])) * 100)
sigma = float(np.std(resid, ddof=1))

forecast = model.forecast(HORIZON)
fidx = pd.date_range(series.index[-1] + pd.offsets.MonthBegin(1), periods=HORIZON, freq="MS")
steps = np.arange(1, HORIZON + 1)
se = sigma * np.sqrt(1 + (steps - 1) * alpha**2)
lower = np.clip(forecast.values - 1.96 * se, 0, None)
upper = forecast.values + 1.96 * se

# ----------------------------------------------------------------- styles
FONT = "Arial"
hdr_fill = PatternFill("solid", fgColor="4E79A7")
fc_fill = PatternFill("solid", fgColor="DCE6F1")
title_font = Font(name=FONT, size=14, bold=True, color="1F3864")
hdr_font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
bold = Font(name=FONT, size=11, bold=True)
normal = Font(name=FONT, size=11)
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ================================================================= Sheet 1: Ringkasan
ws = wb.active
ws.title = "Ringkasan"
ws.sheet_view.showGridLines = False
ws["B2"] = "Hasil Forecast — Single Exponential Smoothing (SES)"
ws["B2"].font = title_font
ws["B3"] = "Dataset: Case 2 FACT  |  Horizon ramalan: 24 bulan (Jan 2025 – Des 2026)"
ws["B3"].font = Font(name=FONT, size=10, italic=True, color="808080")

info = [
    ("Metode", "Single Exponential Smoothing"),
    ("Periode historis", f"{series.index.min():%b %Y} – {series.index.max():%b %Y} ({len(series)} bulan)"),
    ("Alpha (smoothing level)", round(alpha, 4)),
    ("MAPE in-sample", f"{mape:.2f}%"),
    ("Level ramalan (konstan)", round(float(forecast.values[0]))),
    ("Std. deviasi residual (σ)", round(sigma, 1)),
]
r = 5
for k, v in info:
    ws[f"B{r}"] = k; ws[f"B{r}"].font = bold
    ws[f"C{r}"] = v; ws[f"C{r}"].font = normal
    r += 1

ws[f"B{r+1}"] = "Catatan:"
ws[f"B{r+1}"].font = bold
notes = [
    "SES hanya menangkap komponen LEVEL, sehingga ramalan berupa nilai konstan.",
    f"{n_missing} bulan hilang (Jan–Mar 2024) diisi dengan interpolasi linear.",
    "Interval 95% = forecast ± 1.96 × SE; SE membesar seiring horizon (ketidakpastian naik).",
    "Angka forecast dihitung dengan model statsmodels SimpleExpSmoothing (bukan formula Excel).",
]
for i, n in enumerate(notes):
    c = ws[f"B{r+2+i}"]; c.value = f"• {n}"; c.font = Font(name=FONT, size=10, color="595959")

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 45

# ================================================================= Sheet 2: Data & Forecast
ws2 = wb.create_sheet("Data & Forecast")
ws2.sheet_view.showGridLines = False
headers = ["Bulan", "Aktual", "SES Fitted", "Forecast", "Lower 95%", "Upper 95%"]
for j, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=j, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

row = 2
# historis
for dt, act, fit in zip(series.index, series.values, fitted.values):
    ws2.cell(row=row, column=1, value=dt.strftime("%b %Y")).font = normal
    ws2.cell(row=row, column=2, value=float(act)).font = normal
    ws2.cell(row=row, column=3, value=round(float(fit), 1)).font = normal
    for col in range(1, 7):
        ws2.cell(row=row, column=col).border = border
    row += 1
# forecast
for dt, fc, lo, up in zip(fidx, forecast.values, lower, upper):
    ws2.cell(row=row, column=1, value=dt.strftime("%b %Y")).font = normal
    ws2.cell(row=row, column=4, value=round(float(fc))).font = bold
    ws2.cell(row=row, column=5, value=round(float(lo))).font = normal
    ws2.cell(row=row, column=6, value=round(float(up))).font = normal
    for col in range(1, 7):
        cell = ws2.cell(row=row, column=col)
        cell.border = border; cell.fill = fc_fill
    row += 1
last = row - 1

num_fmt = "#,##0"
for col in range(2, 7):
    for rr in range(2, row):
        ws2.cell(row=rr, column=col).number_format = num_fmt
ws2.column_dimensions["A"].width = 12
for col in "BCDEF":
    ws2.column_dimensions[col].width = 13

# ---- native Excel line chart
chart = LineChart()
chart.title = "SES Forecast — Aktual vs Fitted vs Forecast (2 tahun)"
chart.style = 2
chart.height = 10; chart.width = 24
chart.y_axis.title = "Value"; chart.x_axis.title = "Bulan"
cats = Reference(ws2, min_col=1, min_row=2, max_row=last)
for col, name in [(2, "Aktual"), (3, "SES Fitted"), (4, "Forecast"),
                  (5, "Lower 95%"), (6, "Upper 95%")]:
    ref = Reference(ws2, min_col=col, min_row=1, max_row=last)
    chart.add_data(ref, titles_from_data=True)
chart.set_categories(cats)
# style tweaks: forecast series dashed
colors = ["000000", "59A14F", "4E79A7", "BFBFBF", "BFBFBF"]
for s, hexc in zip(chart.series, colors):
    s.graphicalProperties.line.solidFill = hexc
    s.smooth = False
ws2.add_chart(chart, "H2")

# ================================================================= Sheet 3: Grafik (PNG)
ws3 = wb.create_sheet("Grafik")
ws3.sheet_view.showGridLines = False
ws3["B2"] = "Grafik Forecast SES 2 Tahun (dengan interval 95%)"
ws3["B2"].font = title_font
img = XLImage("ses_forecast_2years.png")
img.width = int(img.width * 0.85)
img.height = int(img.height * 0.85)
ws3.add_image(img, "B4")

wb.save("SES_Forecast_Result.xlsx")
print("[ok] SES_Forecast_Result.xlsx tersimpan")
print(f"     alpha={alpha:.4f} mape={mape:.2f}% level={forecast.values[0]:,.0f}")
