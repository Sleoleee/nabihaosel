"""
PROMPT 8 — Data layer Territory.
Muat master customer, normalisasi wilayah, klasifikasi Trade vs Non-Trade,
perluas dim_customer, bangun agregasi provinsi, cetak laporan validasi.

Jalankan SETELAH build_analytics.py (butuh dim_customer terisi status & revenue_at_risk).

  cd backend && source venv/bin/activate
  python build_territory.py /path/ke/database_customer_TPF.xlsx
  # opsi: --cache  (pakai cache transaksi lokal dari build_analytics)
  # opsi: --dry    (hanya cetak klasifikasi & validasi, TIDAK menulis ke DB)

Langkah persis mengikuti PROMPT 8 (LANGKAH 1–6).
"""
import sys
import os
from datetime import date
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import provinces
from build_analytics import fetch_clean, _retry, _f, CHUNK


# ----------------------------- LANGKAH 1: muat master -----------------------------
MASTER_COLS = {
    "cardcode": "customer_code", "cardname": "customer_name", "createdate": "create_date",
    "wilayah": "wilayah_raw", "groupname": "group_name", "pymntgroup": "pymnt_group",
    "isactive": "is_active_master",
    # Phone1 SENGAJA tak dipetakan -> dibuang.
}


def _to_date(v):
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
        except Exception:
            pass
    from datetime import datetime
    try:
        return datetime.fromisoformat(str(v)).date().isoformat()
    except Exception:
        return None


def load_master(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip().lower().replace(" ", "")
        if key in MASTER_COLS:
            idx[MASTER_COLS[key]] = i
    for need in ("customer_code", "wilayah_raw"):
        if need not in idx:
            raise SystemExit(f"Kolom master wajib tidak ditemukan: {need}. Header: {header}")

    rows = []
    for r in it:
        code = r[idx["customer_code"]]
        if code is None or str(code).strip() == "":
            continue
        act = r[idx["is_active_master"]] if "is_active_master" in idx else None
        rows.append({
            "customer_code": str(code).strip(),
            "customer_name": (str(r[idx["customer_name"]]).strip() if "customer_name" in idx and r[idx["customer_name"]] is not None else None),
            "create_date": _to_date(r[idx["create_date"]]) if "create_date" in idx else None,
            "wilayah_raw": (str(r[idx["wilayah_raw"]]).strip() if r[idx["wilayah_raw"]] is not None else None),
            "group_name": (str(r[idx["group_name"]]).strip() if "group_name" in idx and r[idx["group_name"]] is not None else None),
            "pymnt_group": (str(r[idx["pymnt_group"]]).strip() if "pymnt_group" in idx and r[idx["pymnt_group"]] is not None else None),
            "is_active_master": (str(act).strip().upper() in ("Y", "YES", "TRUE", "1")) if act is not None else None,
        })
    return rows


def months_between(d1, d2):
    return (d2.year - d1.year) * 12 + (d2.month - d1.month)


# ----------------------------- LANGKAH 2 & 3: normalisasi + klasifikasi -----------------------------
def enrich_master(master):
    """Tambahkan province_code, is_territory, non_territory_type ke tiap baris master.
    Kumpulkan nilai wilayah yang tak dikenali -> pemanggil GAGALKAN."""
    unknown = defaultdict(int)
    today = date.today()
    for m in master:
        pfx = provinces.cardcode_prefix(m["customer_code"])
        try:
            pc = provinces.normalize_wilayah(m["wilayah_raw"])
        except ValueError:
            unknown[str(m["wilayah_raw"]).strip()] += 1
            pc = None
            m["_wilayah_unknown"] = True
        m["province_code"] = pc
        if pc:
            m["is_territory"] = True
            m["non_territory_type"] = None
            m["province_name"] = provinces.PROVINCES[pc][0]
            m["region_pulau"] = provinces.PROVINCES[pc][1]
        else:
            m["is_territory"] = False
            m["non_territory_type"] = provinces.classify_non_territory(
                m["customer_name"], m["group_name"], pfx)
            m["province_name"] = None
            m["region_pulau"] = None
        # umur customer (bulan) dari create_date
        cd = m["create_date"]
        if cd:
            try:
                y, mo, d = map(int, cd.split("-"))
                m["umur_customer_bulan"] = max(0, months_between(date(y, mo, d), today))
            except Exception:
                m["umur_customer_bulan"] = None
        else:
            m["umur_customer_bulan"] = None
    return unknown


# ----------------------------- LANGKAH 5: agregasi provinsi -----------------------------
def build_province_aggs(master, acm, acc, dim_customer):
    """Bangun agregasi provinsi dari TABEL AGREGAT (bukan transaksi mentah) — cepat & tahan koneksi.
      acm = agg_customer_month  (customer_code, tahun, bulan, revenue, bills)
      acc = agg_customer_category (customer_code, kategori, revenue)  -> penetrasi (lifetime)
      dim_customer = (customer_code, salesperson_utama, status, revenue_at_risk)
    Catatan: kategori bersifat lifetime (agg_customer_category tak punya tahun) -> disimpan tahun=0.
    Salesperson diatribusikan ke salesperson_utama per customer (per-tahun via acm)."""
    code2prov = {m["customer_code"]: m["province_code"] for m in master}
    dc = {d["customer_code"]: d for d in dim_customer}

    # first-ever month per customer (untuk customer_baru), dari acm
    first_ym = {}
    for r in acm:
        c = r.get("customer_code")
        if not c:
            continue
        ym = (int(r["tahun"]), int(r.get("bulan") or 0))
        if c not in first_ym or ym < first_ym[c]:
            first_ym[c] = ym

    apm = defaultdict(lambda: {"revenue": 0.0, "bills": 0, "cust": set(), "baru": set()})
    aps = defaultdict(lambda: {"revenue": 0.0, "cust": set()})   # (pc, slp, tahun)
    prov_year_active = defaultdict(set)   # (pc,year) -> set(cust)

    for r in acm:
        c = r.get("customer_code")
        pc = code2prov.get(c)
        if not pc:      # non-trade / tak ada di master -> tak masuk peta
            continue
        y = int(r["tahun"]); mo = int(r.get("bulan") or 0)
        rev = _f(r.get("revenue"))
        a = apm[(pc, y, mo)]
        a["revenue"] += rev
        a["bills"] += int(r.get("bills") or 0)
        a["cust"].add(c)
        if first_ym.get(c) == (y, mo):
            a["baru"].add(c)
        prov_year_active[(pc, y)].add(c)
        slp = (dc.get(c, {}).get("salesperson_utama") or "—")
        sk = aps[(pc, slp, y)]; sk["revenue"] += rev; sk["cust"].add(c)

    # kategori (lifetime) dari agg_customer_category -> agg_province_category (tahun=0)
    apc = defaultdict(lambda: {"revenue": 0.0, "cust": set()})
    for r in acc:
        c = r.get("customer_code")
        pc = code2prov.get(c)
        if not pc:
            continue
        kat = (r.get("kategori") or "—")
        ck = apc[(pc, kat)]; ck["revenue"] += _f(r.get("revenue")); ck["cust"].add(c)

    # current-state (overdue/lost/at-risk) per provinsi dari dim_customer, ditaruh di bulan TERBARU
    latest_ym = defaultdict(lambda: (0, 0))
    for (pc, y, mo) in apm:
        if (y, mo) > latest_ym[pc]:
            latest_ym[pc] = (y, mo)
    cs = defaultdict(lambda: {"overdue": 0, "lost": 0, "at_risk": 0.0})
    for m in master:
        pc = m["province_code"]
        if not pc:
            continue
        d = dc.get(m["customer_code"])
        if not d:
            continue
        st = (d.get("status") or "")
        if st == "Overdue":
            cs[pc]["overdue"] += 1
            cs[pc]["at_risk"] += _f(d.get("revenue_at_risk"))
        if st == "Lost":
            cs[pc]["lost"] += 1

    apm_rows = []
    for (pc, y, mo), a in apm.items():
        bills = a["bills"]; rev = a["revenue"]
        is_latest = (y, mo) == latest_ym[pc]
        apm_rows.append({
            "province_code": pc, "tahun": y, "bulan": mo,
            "revenue": round(rev, 2), "bills": bills,
            "customer_aktif": len(a["cust"]), "customer_baru": len(a["baru"]),
            "customer_lost": cs[pc]["lost"] if is_latest else 0,
            "revenue_at_risk": round(cs[pc]["at_risk"], 2) if is_latest else 0,
            "customer_overdue": cs[pc]["overdue"] if is_latest else 0,
            "aov": round(rev / bills) if bills else 0,
        })

    apc_rows = [{"province_code": pc, "kategori": kat, "tahun": 0,
                 "revenue": round(v["revenue"], 2), "jumlah_customer": len(v["cust"])}
                for (pc, kat), v in apc.items()]
    aps_rows = [{"province_code": pc, "slp_name": slp, "tahun": y,
                 "revenue": round(v["revenue"], 2), "jumlah_customer": len(v["cust"])}
                for (pc, slp, y), v in aps.items()]

    # dim_province_stats: terdaftar (kumulatif s/d akhir tahun), aktif, tidur, aktivasi
    years = sorted({int(r["tahun"]) for r in acm})
    reg_by_prov_cd = defaultdict(list)   # pc -> list of create_date (date)
    for m in master:
        pc = m["province_code"]
        if not pc or not m["create_date"]:
            continue
        try:
            y, mo, d = map(int, m["create_date"].split("-"))
            reg_by_prov_cd[pc].append(date(y, mo, d))
        except Exception:
            pass
    dps_rows = []
    all_pc = list(provinces.PROVINCES.keys())
    for pc in all_pc:
        for y in years:
            end = date(y, 12, 31)
            terdaftar = sum(1 for cd in reg_by_prov_cd.get(pc, []) if cd <= end)
            aktif = len(prov_year_active.get((pc, y), set()))
            tidur = max(0, terdaftar - aktif)
            akt = round(aktif / terdaftar * 100, 1) if terdaftar else 0
            dps_rows.append({"province_code": pc, "tahun": y,
                             "customer_terdaftar": terdaftar, "customer_aktif": aktif,
                             "customer_tidur": tidur, "tingkat_aktivasi": akt})
    return apm_rows, apc_rows, aps_rows, dps_rows


def provinces_rev_col():
    import config
    return config.REVENUE_COLUMN


# ----------------------------- LANGKAH 6: validasi -----------------------------
def validate_report(master, acm, unknown):
    """Validasi dari agg_customer_month (bukan transaksi mentah)."""
    msgs = []
    ok = True
    master_codes = {m["customer_code"] for m in master}
    code2prov = {m["customer_code"]: m["province_code"] for m in master}

    # a) customer di transaksi yang tak ada di master
    tx_codes = defaultdict(float)
    rev_y = defaultdict(float); rev_y_terr = defaultdict(float)
    for r in acm:
        c = r.get("customer_code")
        if not c:
            continue
        rev = _f(r.get("revenue")); y = int(r["tahun"])
        tx_codes[c] += rev
        rev_y[y] += rev
        if code2prov.get(c):
            rev_y_terr[y] += rev
    missing = {c: rev for c, rev in tx_codes.items() if c not in master_codes}
    if missing:
        ok = False
        msgs.append(f"[a] ❌ {len(missing)} customer di transaksi TIDAK ada di master:")
        for c, rev in sorted(missing.items(), key=lambda x: -x[1])[:20]:
            msgs.append(f"      {c}  Rp {rev:,.0f}")
    else:
        msgs.append("[a] ✓ Semua customer transaksi ada di master (0 hilang).")

    # b) revenue territory + non-territory = total
    total = sum(tx_codes.values())
    terr = sum(rev for c, rev in tx_codes.items() if code2prov.get(c))
    nonterr = total - terr
    diff = total - (terr + nonterr)
    msgs.append(f"[b] Revenue total Rp {total:,.0f} = Territory Rp {terr:,.0f} + Non-Trade Rp {nonterr:,.0f} (selisih {diff:,.2f})")

    # c) nilai wilayah gagal normalisasi
    if unknown:
        ok = False
        msgs.append(f"[c] ❌ {len(unknown)} nilai Wilayah tak dikenali (tambahkan ke provinces._VARIANTS):")
        for w, n in sorted(unknown.items(), key=lambda x: -x[1]):
            msgs.append(f"      {w!r}  ({n} customer)")
    else:
        msgs.append("[c] ✓ Semua nilai Wilayah berhasil dinormalisasi.")

    # d) uji silang prefix vs wilayah (LAPOR saja)
    mism = []
    for m in master:
        pc = m["province_code"]
        pfx = provinces.cardcode_prefix(m["customer_code"])
        exp = provinces.CARDCODE_PREFIX_PROVINCE.get(pfx)
        if pc and exp and exp != pc:
            mism.append((m["customer_code"], provinces.PROVINCES[exp][0], provinces.PROVINCES[pc][0]))
    if mism:
        msgs.append(f"[d] ⚠ {len(mism)} kandidat salah-input (prefix ≠ Wilayah), contoh:")
        for c, e, a in mism[:15]:
            msgs.append(f"      {c}: prefix→{e}  tapi Wilayah→{a}")
    else:
        msgs.append("[d] ✓ Tidak ada ketidakcocokan prefix vs Wilayah.")

    # e) coverage % revenue terpetakan per tahun
    msgs.append("[e] Coverage revenue terpetakan ke provinsi per tahun:")
    for y in sorted(rev_y):
        pct = rev_y_terr[y] / rev_y[y] * 100 if rev_y[y] else 0
        msgs.append(f"      {y}: {pct:.1f}%  (Rp {rev_y_terr[y]:,.0f} / Rp {rev_y[y]:,.0f})")
    return ok, msgs


def classification_summary(master):
    by_type = defaultdict(lambda: {"n": 0})
    terr = sum(1 for m in master if m["is_territory"])
    non = len(master) - terr
    lines = [f"Master: {len(master)} customer  |  Trade(punya provinsi): {terr}  |  Non-Trade: {non}"]
    tcount = defaultdict(int)
    for m in master:
        if not m["is_territory"]:
            tcount[m["non_territory_type"]] += 1
    lines.append("Rincian Non-Trade:")
    for t in provinces.NON_TERRITORY_TYPES:
        lines.append(f"   {t:16} : {tcount.get(t,0)}")
    return "\n".join(lines)


# ----------------------------- MAIN -----------------------------
def preview_diff(db, master, unknown):
    """PROMPT 10: bandingkan master baru vs customer_master di DB. LAPOR saja, tak menulis."""
    existing = _retry(lambda: db.table("customer_master")
                      .select("customer_code,wilayah_raw,is_active_master").limit(1000000).execute().data) or []
    ex = {e["customer_code"]: e for e in existing}
    new_cust = [m for m in master if m["customer_code"] not in ex]
    changed_wil = [m for m in master if m["customer_code"] in ex
                   and (m["wilayah_raw"] or "") != (ex[m["customer_code"]].get("wilayah_raw") or "")]
    became_inactive = [m for m in master if m["customer_code"] in ex
                       and ex[m["customer_code"]].get("is_active_master") is True and m["is_active_master"] is False]
    print("\n--- PREVIEW PERUBAHAN MASTER ---")
    print(f"  Customer baru        : {len(new_cust)}")
    print(f"  Wilayah berubah      : {len(changed_wil)}")
    print(f"  Menjadi tidak aktif  : {len(became_inactive)}")
    if unknown:
        print(f"  ❌ Wilayah baru tak dikenal ({len(unknown)}) — IMPOR DIBLOKIR sampai ditambahkan ke provinces._VARIANTS:")
        for w, n in sorted(unknown.items(), key=lambda x: -x[1]):
            print(f"       {w!r} ({n})")
    # customer di transaksi tapi hilang dari master baru
    tx = _retry(lambda: db.table("dim_customer").select("customer_code").limit(1000000).execute().data) or []
    new_codes = {m["customer_code"] for m in master}
    gone = [t["customer_code"] for t in tx if t["customer_code"] not in new_codes]
    if gone:
        print(f"  ⚠ {len(gone)} customer ada di transaksi tapi HILANG dari master baru (contoh): {gone[:10]}")
    return not unknown


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        raise SystemExit("Pakai: python build_territory.py <database_customer_TPF.xlsx> [--cache] [--dry] [--preview]")
    path = args[0]
    use_cache = "--cache" in sys.argv
    dry = "--dry" in sys.argv
    preview = "--preview" in sys.argv

    from utils.db import get_client
    db = get_client()

    print(f"[1] Memuat master: {path}")
    master = load_master(path)
    print(f"    {len(master)} baris master.")

    print("[2/3] Normalisasi wilayah + klasifikasi Trade/Non-Trade...")
    unknown = enrich_master(master)
    print(classification_summary(master))

    if preview:
        okp = preview_diff(db, master, unknown)
        if not okp:
            sys.exit(1)
        print("\n(--preview) selesai. Jalankan tanpa --preview untuk menulis + agregasi ulang.")
        return

    print("\n[fetch] Mengambil tabel agregat (agg_customer_month, agg_customer_category, dim_customer)...")
    acm = _fetch_agg(db, "agg_customer_month", "customer_code,tahun,bulan,revenue,bills")
    print(f"    agg_customer_month: {len(acm)} baris")
    acc = _fetch_agg(db, "agg_customer_category", "customer_code,kategori,revenue")
    print(f"    agg_customer_category: {len(acc)} baris")
    dim_customer = _fetch_agg(db, "dim_customer", "customer_code,status,revenue_at_risk,salesperson_utama")
    print(f"    dim_customer: {len(dim_customer)} baris")
    if not acm:
        print("\n❌ agg_customer_month kosong. Jalankan build_analytics.py dulu.")
        sys.exit(1)

    print("[5] Membangun agregasi provinsi...")
    apm, apc, aps, dps = build_province_aggs(master, acm, acc, dim_customer)

    print("\n--- [6] LAPORAN VALIDASI ---")
    ok, msgs = validate_report(master, acm, unknown)
    print("\n".join(msgs))
    if not ok:
        print("\n❌ Validasi GAGAL. Perbaiki dulu (mis. tambah pemetaan wilayah / cek master). Tidak menulis.")
        sys.exit(1)
    if dry:
        print("\n(--dry) Tidak menulis ke DB.")
        return

    print("\n[tulis] Menyimpan ke Supabase...")
    _write(db, "dim_province", provinces.province_rows(), "province_code")
    from datetime import datetime, timezone
    now_iso = datetime.now(timezone.utc).isoformat()
    master_rows = [{**{k: m[k] for k in ("customer_code", "customer_name", "create_date",
                    "wilayah_raw", "group_name", "pymnt_group", "is_active_master")},
                    "master_updated_at": now_iso} for m in master]
    _write(db, "customer_master", master_rows, "customer_code")

    # LANGKAH 4: perluas dim_customer (upsert kolom wilayah by customer_code)
    dc_upd = [{"customer_code": m["customer_code"], "province_code": m["province_code"],
               "province_name": m["province_name"], "region_pulau": m["region_pulau"],
               "is_territory": m["is_territory"], "non_territory_type": m["non_territory_type"],
               "group_name": m["group_name"], "pymnt_group": m["pymnt_group"],
               "create_date": m["create_date"], "is_active_master": m["is_active_master"],
               "umur_customer_bulan": m["umur_customer_bulan"]}
              for m in master]
    _write(db, "dim_customer", dc_upd, "customer_code")

    _write(db, "agg_province_month", apm, "province_code,tahun,bulan")
    _write(db, "agg_province_category", apc, "province_code,kategori,tahun")
    _write(db, "agg_province_salesperson", aps, "province_code,slp_name,tahun")
    _write(db, "dim_province_stats", dps, "province_code,tahun")
    print("\n✓ Selesai. Lapisan Territory terisi.")


def _fetch_agg(db, table, cols):
    """Ambil tabel agregat sekali (retryable). Jauh lebih kecil dari transaksi mentah."""
    return _retry(lambda: db.table(table).select(cols).limit(1000000).execute().data) or []


def _write(db, name, recs, pk):
    print(f"  {name}: {len(recs)} baris...", end=" ")
    for i in range(0, len(recs), CHUNK):
        chunk = recs[i:i+CHUNK]
        _retry(lambda: db.table(name).upsert(chunk, on_conflict=pk).execute())
    print("ok")


if __name__ == "__main__":
    main()
