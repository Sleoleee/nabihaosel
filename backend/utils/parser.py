import openpyxl
import re
from datetime import datetime
from typing import Dict, List, Tuple


COLUMN_MAP = {
    "Document Status": "document_status",
    "Canceled": "canceled",
    "Document Number": "document_number",
    "Posting Date": "posting_date",
    "Due Date": "due_date",
    "Customer/Vendor_Code": "customer_code",
    "Customer/Vendor_Name": "customer_name",
    "Item_No.": "item_no",
    "Item/Service_Description": "item_description",
    "Kategori": "kategori",
    "Warehouse Code": "warehouse_code",
    "Quantity": "quantity",
    "Unit": "unit",
    "harga_awal": "harga_awal",
    "Disc % Per Row": "disc_per_row",
    "harga_jual": "harga_jual",
    "Row Total": "row_total",
    "Disc % For Document": "disc_for_document",
    "New_Row_Total": "new_row_total",
    "SlpName": "slp_name",
    "Branch": "branch",
    "Status_payment": "status_payment",
    # Varian nama kolom pada file AR (2026)
    "Due_Date": "due_date",
    "Disc% For Document": "disc_for_document",
}

NUMERIC_COLS = {"harga_awal", "disc_per_row", "harga_jual", "row_total",
                "disc_for_document", "new_row_total"}
# quantity kolom DB bertipe INTEGER -> harus dikirim sebagai int, bukan 1.0
INT_COLS = {"document_number", "quantity"}
DATE_COLS = {"posting_date", "due_date"}
OUTPUT_COLS = [
    "document_number", "posting_date", "due_date", "customer_code",
    "customer_name", "item_no", "item_description", "kategori",
    "warehouse_code", "quantity", "unit", "harga_awal", "disc_per_row",
    "harga_jual", "row_total", "disc_for_document", "new_row_total",
    "slp_name", "branch", "status_payment", "year",
]


def _to_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _to_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if hasattr(val, 'isoformat'):
        return val.isoformat()
    try:
        return datetime.fromisoformat(str(val)).date().isoformat()
    except Exception:
        return None


def _fixed_year_from_name(sheet_name):
    """Tahun dari nama sheet 'sales detail YYYY', atau None (tahun diambil per baris)."""
    m = re.match(r"sales detail (\d{4})", sheet_name, re.IGNORECASE)
    return int(m.group(1)) if m else None


def parse_excel(file_bytes: bytes) -> Dict[int, Tuple[List[dict], int]]:
    """
    Returns dict: {year: (records, skipped_count)}
    Streams rows without building a full DataFrame to minimize memory usage.
    Sebuah sheet dianggap sheet transaksi bila header-nya memuat kolom inti
    (Posting Date + New_Row_Total) — jadi robust untuk nama sheet apa pun
    (sales detail YYYY, AR, labeled_data, Sheet1, ...). Tahun diambil dari nama
    sheet bila cocok "sales detail YYYY", selain itu dari Posting Date per baris.
    Sheet ringkasan (mis. subkategori_summary) & master (BP) otomatis dilewati.
    """
    wb = openpyxl.load_workbook(filename=file_bytes, read_only=True, data_only=True)
    acc = {}  # year -> [records_list, skipped_count]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        # Build index map: excel column name -> db column name -> position
        col_index = {}
        for i, h in enumerate(header_row):
            if h is None:
                continue
            h_str = str(h).strip()
            db_name = COLUMN_MAP.get(h_str)
            if db_name is not None and db_name not in col_index:
                col_index[db_name] = i

        # Hanya proses sheet transaksi (punya kolom inti); lainnya dilewati
        if "posting_date" not in col_index or "new_row_total" not in col_index:
            continue
        fixed_year = _fixed_year_from_name(sheet_name)

        def get_cell(row, col):
            idx = col_index.get(col)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in rows_iter:
            posting = _to_date(get_cell(row, "posting_date"))
            # Tentukan tahun: dari nama sheet, atau dari Posting Date (sheet AR)
            year = fixed_year if fixed_year is not None else (
                int(posting[:4]) if posting and len(posting) >= 4 else None)
            if year is None:
                continue  # tak bisa ditentukan tahunnya, lewati diam-diam
            bucket = acc.setdefault(year, [[], 0])

            # Filter: skip canceled rows
            canceled_val = get_cell(row, "canceled")
            if canceled_val is not None and str(canceled_val).strip().upper() != "N":
                bucket[1] += 1
                continue

            # Filter: skip zero/negative new_row_total
            nrt = _to_float(get_cell(row, "new_row_total"))
            if nrt <= 0:
                bucket[1] += 1
                continue

            rec = {"year": year}
            for col in OUTPUT_COLS:
                if col == "year":
                    continue
                if col == "posting_date":
                    rec[col] = posting
                    continue
                val = get_cell(row, col)
                if col in INT_COLS:
                    rec[col] = _to_int(val)
                elif col in NUMERIC_COLS:
                    rec[col] = _to_float(val)
                elif col in DATE_COLS:
                    rec[col] = _to_date(val)
                else:
                    rec[col] = str(val).strip() if val is not None else None

            bucket[0].append(rec)

    wb.close()
    return {y: (v[0], v[1]) for y, v in acc.items()}
